"""
ASP Report Generator — 프로그램 2
===================================
raw_asp_data.csv를 읽어서 Excel 보고서 생성
- 요약 시트 + Molecule별 시트
- 제품별 ASP 분기 추이 표 + 꺾은선/막대 그래프
- IRA qualifying 배지, Samsung Bioepis 하이라이트
"""
import os, sys, json, io
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from tkinter import filedialog, Tk

# ── 임상 단위 기준 ──────────────────────────────────────────
CLINICAL_UNITS = {
    "Infliximab":   {"mult": 10,  "unit": "100mg/vial"},
    "Ranibizumab":  {"mult": 5,   "unit": "0.5mg/inj"},
    "Trastuzumab":  {"mult": 42,  "unit": "420mg/vial"},
    "Denosumab":    {"mult": 120, "unit": "120mg/vial"},
    "Bevacizumab":  {"mult": 40,  "unit": "400mg/vial"},
    "Rituximab":    {"mult": 50,  "unit": "500mg/vial"},
    "Filgrastim":   {"mult": 300, "unit": "300mcg/vial"},
    "Pegfilgrastim":{"mult": 12,  "unit": "6mg/syringe"},
    "Epoetin alfa": {"mult": 40,  "unit": "40,000u/vial"},
    "Tocilizumab":  {"mult": 400, "unit": "400mg/vial"},
}

# ── 색상 팔레트 ─────────────────────────────────────────────
CLR = {
    "header_sb":    "003087",  # 진남색 (SB molecule)
    "header_nosb":  "37474F",  # 다크그레이 (시장 검증용)
    "orig":         "C62828",  # 빨강 (Originator)
    "sb":           "1565C0",  # 파랑 (Samsung Bioepis)
    "other":        "607D8B",  # 회색 (기타 biosimilar)
    "ira":          "6A1B9A",  # 보라 (IRA qualifying)
    "row_orig":     "FFEBEE",
    "row_sb":       "E3F0FF",
    "row_other":    "FFFFFF",
    "row_alt":      "F5F5F5",
    "summary_hdr":  "1565C0",
    "q_header":     "E8EAF6",
}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── 스타일 헬퍼 ─────────────────────────────────────────────
def cell_style(ws, row, col, value=None, bold=False, color=None,
               bg=None, align="left", num_fmt=None, border=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=color or "000000", size=size)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = BORDER
    return c


def merge_header(ws, row, c1, c2, text, bg, color="FFFFFF", size=11):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font = Font(name="Arial", bold=True, color=color, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER


# ── 그래프 생성 ─────────────────────────────────────────────
def make_charts(df_mol: pd.DataFrame, mol_name: str,
                quarters: list, mult: int, unit: str) -> tuple:
    """꺾은선 + 막대 그래프 PNG bytes 반환"""

    # 제품별 ASP 피벗
    pivot = df_mol.pivot_table(
        index="quarter", columns="brand",
        values="asp_clinical", aggfunc="first"
    ).reindex(quarters)

    brands  = pivot.columns.tolist()
    orig_b  = df_mol[df_mol["product_type"]=="Originator"]["brand"].iloc[0] if len(df_mol[df_mol["product_type"]=="Originator"]) else None
    sb_bs   = df_mol[df_mol["is_sb"]==True]["brand"].unique().tolist()

    def get_color(b):
        if b == orig_b: return "#C62828"
        if b in sb_bs:  return "#1565C0"
        others = ["#607D8B","#78909C","#90A4AE","#546E7A","#455A64","#37474F"]
        idx = [x for x in brands if x != orig_b and x not in sb_bs].index(b) if b in [x for x in brands if x != orig_b and x not in sb_bs] else 0
        return others[idx % len(others)]

    # ─ 꺾은선 그래프 ─
    fig1, ax1 = plt.subplots(figsize=(13, 5))
    fig1.patch.set_facecolor("#F8F9FA")
    ax1.set_facecolor("#F8F9FA")

    lines_meta = []
    for b in brands:
        vals = pivot[b].values
        xs   = [i for i, v in enumerate(vals) if not pd.isna(v)]
        ys   = [vals[i] for i in xs]
        if not xs: continue
        color = get_color(b)
        lw = 3 if b == orig_b else (2.5 if b in sb_bs else 1.5)
        ls = "-" if b == orig_b else ("--" if b in sb_bs else ":")
        ms = 8 if b == orig_b else (7 if b in sb_bs else 5)
        ax1.plot(xs, ys, marker="o", lw=lw, ls=ls, color=color, ms=ms, zorder=3)
        lines_meta.append((xs[-1], ys[-1], b, color, b==orig_b or b in sb_bs))

    # 라벨 겹침 방지
    lines_meta.sort(key=lambda x: x[1], reverse=True)
    placed = []
    y_vals = [m[1] for m in lines_meta]
    gap = (max(y_vals)-min(y_vals))*0.09 if len(y_vals)>1 and max(y_vals)!=min(y_vals) else 20
    for (xe, ye, txt, color, bold) in lines_meta:
        yp = ye
        for py in placed:
            if abs(yp-py) < gap: yp = py+gap
        placed.append(yp)
        ax1.annotate(txt, xy=(xe, ye), xytext=(xe+0.2, yp),
                     fontsize=7, color=color,
                     fontweight="bold" if bold else "normal", va="center",
                     bbox=dict(boxstyle="round,pad=0.2", fc="white", ec=color, alpha=0.85, lw=0.8),
                     arrowprops=dict(arrowstyle="-", color=color, lw=0.6, alpha=0.5) if abs(yp-ye)>2 else None)

    ax1.set_xticks(range(len(quarters)))
    ax1.set_xticklabels(quarters, fontsize=7, rotation=30, ha="right")
    ax1.set_xlim(-0.3, len(quarters)-1+3.5)
    ax1.set_title(f"{mol_name} — ASP 분기별 추이 ({unit})", fontsize=11, fontweight="bold", pad=8)
    ax1.set_ylabel(f"ASP (USD / {unit})", fontsize=9)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:,.0f}"))
    ax1.grid(axis="y", ls="--", alpha=0.35)
    ax1.spines[["top","right"]].set_visible(False)
    plt.tight_layout()
    buf1 = io.BytesIO()
    plt.savefig(buf1, format="png", dpi=130, bbox_inches="tight")
    plt.close(fig1)
    buf1.seek(0)

    return buf1


# ── Molecule 시트 생성 ──────────────────────────────────────
def build_mol_sheet(wb: Workbook, df_mol: pd.DataFrame,
                    mol_name: str, quarters: list,
                    mult: int, unit: str, has_sb: bool):

    ws = wb.create_sheet(mol_name[:31])
    ws.freeze_panes = "D3"

    # 헤더색
    hdr_bg = CLR["header_sb"] if has_sb else CLR["header_nosb"]

    # ── 타이틀 ──
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+len(quarters))
    c = ws.cell(row=1, column=1,
                value=f"{mol_name}  |  ASP 분기별 추이  |  기준: {unit}"
                      + ("  ★ Samsung Bioepis 제품 포함" if has_sb else "  (시장 검증용)"))
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.fill = PatternFill("solid", fgColor=hdr_bg)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── 컬럼 헤더 ──
    ws.row_dimensions[2].height = 22
    headers = ["Brand", "INN Suffix / HCPCS", "Company"] + quarters
    col_widths = [22, 26, 22] + [13]*len(quarters)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        bg = CLR["q_header"] if ci > 3 else "E8EAF6"
        cell_style(ws, 2, ci, h, bold=True, bg=bg, align="center", size=9)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── 데이터 행 ──
    brands_order = (
        df_mol[df_mol["product_type"]=="Originator"]["brand"].unique().tolist() +
        df_mol[(df_mol["product_type"]=="Biosimilar") & (df_mol["is_sb"]==True)]["brand"].unique().tolist() +
        df_mol[(df_mol["product_type"]=="Biosimilar") & (df_mol["is_sb"]==False)]["brand"].unique().tolist()
    )

    pivot = df_mol.pivot_table(
        index="brand", columns="quarter",
        values=["asp_clinical","payment_limit_clinical","addon_pct","ira_qualifying","hcpcs_code","suffix","company","product_type","is_sb"],
        aggfunc="first"
    )

    row_idx = 3
    for brand in brands_order:
        if brand not in df_mol["brand"].values: continue
        bdf = df_mol[df_mol["brand"]==brand].iloc[0]
        ptype = bdf["product_type"]
        is_sb = bdf["is_sb"]

        if ptype == "Originator":
            bg = CLR["row_orig"]; tag = " [Originator]"; col = CLR["orig"]
        elif is_sb:
            bg = CLR["row_sb"];   tag = " [Samsung Bioepis]"; col = CLR["sb"]
        else:
            bg = CLR["row_other"] if row_idx % 2 == 0 else CLR["row_alt"]
            tag = ""; col = "000000"

        ws.row_dimensions[row_idx].height = 32
        suffix_str = bdf.get("suffix","") or ""
        cell_style(ws, row_idx, 1, brand+tag, bold=(ptype=="Originator" or is_sb),
                   color=col, bg=bg, size=9)
        cell_style(ws, row_idx, 2, suffix_str, bg=bg, size=8, color="555555")
        cell_style(ws, row_idx, 3, bdf["company"], bg=bg, size=8, color="555555")

        prev_asp = None
        for qi, q in enumerate(quarters):
            ci = 4 + qi
            qdf = df_mol[(df_mol["brand"]==brand) & (df_mol["quarter"]==q)]
            if qdf.empty:
                cell_style(ws, row_idx, ci, "—", bg=bg, align="center", color="CCCCCC", size=9)
            else:
                r = qdf.iloc[0]
                asp = r["asp_clinical"]
                pl  = r["payment_limit_clinical"]
                addon = r["addon_pct"]
                ira_q = r["ira_qualifying"]
                hcpcs = r["hcpcs_code"]

                # 전분기 대비
                arrow = ""
                if prev_asp is not None and not pd.isna(asp) and not pd.isna(prev_asp):
                    pct = (asp - prev_asp) / prev_asp * 100
                    if pct > 0.5:   arrow = f" ▲{pct:.1f}%"
                    elif pct < -0.5:arrow = f" ▼{abs(pct):.1f}%"

                ira_mark = " ★IRA" if ira_q else ""
                val_str = f"${asp:,.2f}{ira_mark}{arrow}\nPL:${pl:,.2f} | {hcpcs}"

                c = ws.cell(row=row_idx, column=ci, value=val_str)
                asp_color = CLR["ira"] if ira_q else col
                c.font = Font(name="Arial", size=8, color=asp_color,
                              bold=(ptype=="Originator" or is_sb))
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                c.border = BORDER
                prev_asp = asp if not pd.isna(asp) else prev_asp

        row_idx += 1

    # ── 범례 ──
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    leg = ws.cell(row=row_idx, column=1,
                  value="범례:  ★IRA = IRA qualifying (+8% add-on 적용)  |  ▲▼ = 전분기 대비 변화율  |  PL = Payment Limit")
    leg.font = Font(name="Arial", size=8, color="666666", italic=True)
    leg.alignment = Alignment(horizontal="left", vertical="center")

    # ── 그래프 삽입 ──
    row_idx += 2
    try:
        buf1 = make_charts(df_mol, mol_name, quarters, mult, unit)
        img1 = XLImage(buf1)
        img1.width, img1.height = 800, 340
        ws.add_image(img1, f"A{row_idx}")
    except Exception as e:
        ws.cell(row=row_idx, column=1, value=f"그래프 생성 오류: {e}")


# ── 요약 시트 ───────────────────────────────────────────────
def build_summary_sheet(wb: Workbook, df: pd.DataFrame, quarters: list):
    ws = wb.active
    ws.title = "📊 요약"
    ws.freeze_panes = "A4"

    # 타이틀
    ws.row_dimensions[1].height = 35
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws.cell(row=1, column=1,
                value=f"Biosimilar ASP Monitor — Samsung Bioepis Market Report 검증"
                      f"  |  Data: CMS Medicare Part B  |  생성: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor="003087")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # 서브 헤더
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    c2 = ws.cell(row=2, column=1,
                 value=f"수집 분기: {quarters[0]} ~ {quarters[-1]}  |  총 {len(quarters)}개 분기  |  "
                       f"Molecule: {df['molecule'].nunique()}개  |  제품: {df['brand'].nunique()}개")
    c2.font = Font(name="Arial", size=9, color="444444")
    c2.fill = PatternFill("solid", fgColor="EEF2F7")
    c2.alignment = Alignment(horizontal="left", vertical="center")

    # 컬럼 헤더
    ws.row_dimensions[3].height = 22
    hdrs = ["Molecule", "Product Type", "Brand", "Company",
            "최근 ASP", "전분기 대비", "IRA Qualifying", "비고"]
    widths = [18, 12, 22, 22, 14, 12, 14, 20]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell_style(ws, 3, ci, h, bold=True, bg=CLR["summary_hdr"],
                   color="FFFFFF", align="center", size=9)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 최근 분기 데이터
    latest_q = quarters[-1]
    prev_q   = quarters[-2] if len(quarters) >= 2 else None

    df_latest = df[df["quarter"]==latest_q].copy()
    df_prev   = df[df["quarter"]==prev_q].copy() if prev_q else pd.DataFrame()

    row_idx = 4
    for mol_name in df["molecule"].unique():
        df_m = df_latest[df_latest["molecule"]==mol_name]
        if df_m.empty: continue

        brands_order = (
            df_m[df_m["product_type"]=="Originator"]["brand"].tolist() +
            df_m[(df_m["product_type"]=="Biosimilar") & (df_m["is_sb"]==True)]["brand"].tolist() +
            df_m[(df_m["product_type"]=="Biosimilar") & (df_m["is_sb"]==False)]["brand"].tolist()
        )

        for brand in brands_order:
            r = df_m[df_m["brand"]==brand]
            if r.empty: continue
            r = r.iloc[0]

            ptype = r["product_type"]
            is_sb = r["is_sb"]
            ira_q = r["ira_qualifying"]

            if ptype == "Originator":
                bg = CLR["row_orig"]; col = CLR["orig"]
            elif is_sb:
                bg = CLR["row_sb"];   col = CLR["sb"]
            else:
                bg = CLR["row_other"] if row_idx%2==0 else CLR["row_alt"]
                col = "000000"

            ws.row_dimensions[row_idx].height = 20

            # 전분기 대비
            qoc = "—"
            if not df_prev.empty:
                prev_r = df_prev[(df_prev["molecule"]==mol_name) & (df_prev["brand"]==brand)]
                if not prev_r.empty:
                    pv = prev_r.iloc[0]["asp_clinical"]
                    cv = r["asp_clinical"]
                    if not pd.isna(pv) and not pd.isna(cv) and pv != 0:
                        pct = (cv-pv)/pv*100
                        qoc = f"▲{pct:.1f}%" if pct>0.5 else (f"▼{abs(pct):.1f}%" if pct<-0.5 else "—")

            ira_txt = "✅ +8% IRA" if ira_q else "+6%"
            note    = ""
            if ptype=="Originator": note = "Originator"
            elif is_sb:             note = "★ Samsung Bioepis"

            vals = [mol_name, ptype, brand, r["company"],
                    f"${r['asp_clinical']:,.2f}", qoc, ira_txt, note]
            for ci, v in enumerate(vals, 1):
                bold = ptype=="Originator" or is_sb
                c = cell_style(ws, row_idx, ci, v, bold=bold, bg=bg,
                               color=col if ci in [1,3,5] else "000000",
                               align="right" if ci==5 else "center" if ci in [6,7] else "left",
                               size=9)
            row_idx += 1

        # 구분선
        for ci in range(1, 9):
            ws.cell(row=row_idx, column=ci).fill = PatternFill("solid", fgColor="E0E0E0")
        ws.row_dimensions[row_idx].height = 4
        row_idx += 1


# ── 저장 경로 설정 ──────────────────────────────────────────
def get_save_path() -> str:
    path_file = "output_config.json"
    saved_dir = ""

    if os.path.exists(path_file):
        with open(path_file, "r") as f:
            cfg = json.load(f)
            saved_dir = os.path.dirname(cfg.get("output_path",""))

    print("\n  결과 Excel 저장 위치를 선택하세요.")
    print("  1. 팝업 창으로 폴더 선택")
    print("  2. 직접 경로 입력")
    print("  3. 현재 폴더 (output/)")

    choice = input("  선택: ").strip()

    if choice == "1":
        root = Tk(); root.withdraw(); root.attributes("-topmost", True)
        folder = filedialog.askdirectory(
            title="저장 폴더 선택",
            initialdir=saved_dir or os.path.expanduser("~")
        )
        root.destroy()
        if not folder:
            folder = "output"
    elif choice == "2":
        folder = input("  폴더 경로: ").strip() or "output"
    else:
        folder = "output"

    os.makedirs(folder, exist_ok=True)
    ts    = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
    fpath = os.path.join(folder, f"Biosimilar_ASP_Report_{ts}.xlsx")

    # 경로 저장
    with open(path_file, "w") as f:
        json.dump({"output_path": fpath}, f)

    return fpath


# ── CSV 파일 선택 ───────────────────────────────────────────
def get_csv_path() -> str:
    path_file = "output_config.json"
    default   = os.path.join("output", "raw_asp_data.csv")

    if os.path.exists(path_file):
        with open(path_file, "r") as f:
            cfg = json.load(f)
            saved = cfg.get("output_path","")
            guess = os.path.join(os.path.dirname(saved), "raw_asp_data.csv")
            if os.path.exists(guess):
                default = guess

    print(f"\n  CSV 파일 경로")
    print(f"  1. 팝업 창으로 선택")
    print(f"  2. 직접 입력")
    print(f"  3. 기본 경로 사용 ({default})")

    choice = input("  선택: ").strip()

    if choice == "1":
        root = Tk(); root.withdraw(); root.attributes("-topmost", True)
        path = filedialog.askopenfilename(
            title="raw_asp_data.csv 선택",
            filetypes=[("CSV files","*.csv"),("All files","*.*")],
            initialdir=os.path.dirname(default)
        )
        root.destroy()
        return path or default
    elif choice == "2":
        return input("  CSV 경로: ").strip() or default
    else:
        return default


# ── 메인 ────────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  ASP Report Generator — GitHub Actions 모드")
    print("="*55)

    csv_path  = "data/asp_data.csv"
    save_path = "data/asp_report.xlsx"

    if not os.path.exists(csv_path):
        print(f"❌ 파일 없음: {csv_path}")
        return

    print(f"✅ CSV 로드 중...")
    df = pd.read_csv(csv_path, dtype={"is_sb": str, "ira_qualifying": str})
    df["is_sb"]          = df["is_sb"].map({"True":True,"False":False}).fillna(False)
    df["ira_qualifying"] = df["ira_qualifying"].map({"True":True,"False":False}).fillna(False)
    df["asp_per_unit"]   = pd.to_numeric(df["asp_per_unit"], errors="coerce")
    df["payment_limit"]  = pd.to_numeric(df["payment_limit"], errors="coerce")

    def get_mult(mol): return CLINICAL_UNITS.get(mol, {}).get("mult", 1)
    def get_unit(mol): return CLINICAL_UNITS.get(mol, {}).get("unit", "per unit")

    df["asp_clinical"]           = df.apply(lambda r: r["asp_per_unit"]*get_mult(r["molecule"]), axis=1)
    df["payment_limit_clinical"] = df.apply(lambda r: r["payment_limit"]*get_mult(r["molecule"]), axis=1)

    quarters  = df["quarter"].unique().tolist()
    molecules = df["molecule"].unique().tolist()
    sb_molecules = df[df["is_sb"]==True]["molecule"].unique().tolist()

    print(f"  분기: {len(quarters)}개 | Molecule: {len(molecules)}개 | 총 {len(df)}행")

    print(f"\n  Excel 생성 중...\n")
    os.makedirs("data", exist_ok=True)
    wb = Workbook()

    build_summary_sheet(wb, df, quarters)
    print(f"  ✅ 요약 시트 완료")

    for mol_name in molecules:
        df_mol = df[df["molecule"]==mol_name].copy()
        mult   = get_mult(mol_name)
        unit   = get_unit(mol_name)
        has_sb = mol_name in sb_molecules
        build_mol_sheet(wb, df_mol, mol_name, quarters, mult, unit, has_sb)
        print(f"  ✅ {mol_name} 시트 완료")

    wb.save(save_path)
    print(f"\n✅ 저장 완료: {save_path}")


if __name__ == "__main__":
    main()
